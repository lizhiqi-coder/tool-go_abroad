# coding=utf-8
import os
import sys
from xml.dom import minidom
from xml.dom.minidom import parse
import openpyxl
from openpyxl.writer.excel import ExcelWriter

CN_map = {}
EN_map = {}
CN_key_list = []
EN_key_list = []

FILE_PATH_XLSX = 'strings.xlsx'
FILE_PATH_XML = 'strings.xml'
GENERATE_ANDROID_STRINGS_FILE = "strings_en.xml"

ANDROID_KEY_COLUMN_IN_XLSX = 'B'
CN_VALUE_COLUMN_IN_XLSX = 'C'
EN_VALUE_COLUMN_IN_XLSX = 'D'

FIRST_LINE_IN_XLSX = '1'
LAST_LINE_IN_XLSX = '1000'

CONTENT_RANGE_IN_XLSX = ANDROID_KEY_COLUMN_IN_XLSX + FIRST_LINE_IN_XLSX + ':' + EN_VALUE_COLUMN_IN_XLSX + LAST_LINE_IN_XLSX

TITLE_ANDROID_KEY = 'android_key'
TITLE_CN_STRING_VALUE = 'CN_String_value'
TITLE_EN_STRING_VALUE = 'EN_String_value'


def get_content_range_in_xlsx(row_num):
    LAST_LINE_IN_XLSX = str(row_num)
    CONTENT_RANGE_IN_XLSX = ANDROID_KEY_COLUMN_IN_XLSX + FIRST_LINE_IN_XLSX + ':' + EN_VALUE_COLUMN_IN_XLSX + LAST_LINE_IN_XLSX
    return CONTENT_RANGE_IN_XLSX


# 读取已翻译过的xlsx文件，返回中文字典，英文字典，中文顺序关键词列表，英文顺序关键词列表
def read_xlsx_file(file_path):
    global CN_map, EN_map, CN_key_list, EN_key_list
    cleanBuffer()
    wb = openpyxl.load_workbook(filename=file_path, use_iterators=True)
    sheet_names = wb.get_sheet_names()
    head_sheet = wb.get_sheet_by_name(sheet_names[0])
    # LAST_LINE_IN_XLSX = head_sheet.max_row
    for row in head_sheet.iter_rows(get_content_range_in_xlsx(head_sheet.max_row)):
        key = row[0].value

        if (key == None or key == TITLE_ANDROID_KEY):
            continue
        # print  'CN_key_list', CN_key_list
        CN_key_list.append(key)
        EN_key_list.append(key)
        CN_value = row[1].value
        EN_value = row[2].value
        if CN_value != None:
            CN_map[key] = CN_value.encode('utf-8')
        else:
            CN_map[key] = CN_value

        if EN_value != None:
            EN_map[key] = EN_value.encode('utf-8')
        else:
            EN_map[key] = EN_value

    return CN_map, EN_map, CN_key_list, EN_key_list


def write_to_xlsx_file(content, keys, file_path):
    if os.path.exists(file_path) == False:
        createFile(file_path)

    wb = openpyxl.Workbook()
    ew = ExcelWriter(workbook=wb)
    wsheet = wb.worksheets[0]

    is_close = raw_input('please be sure to close your xlsx file (y/n)?')
    if is_close != 'y':
        print '未关闭文件，退出程序'
        return

    rows = [row for row in wsheet.iter_rows(get_content_range_in_xlsx(keys.__len__() + 1))]
    rows[0][0].value = TITLE_ANDROID_KEY
    rows[0][1].value = TITLE_CN_STRING_VALUE
    rows[0][2].value = TITLE_EN_STRING_VALUE

    for i in range(0, keys.__len__()):
        rows[i + 1][0].value = keys[i]
        rows[i + 1][1].value = content[keys[i]]
        continue
        # print rows[i + 1][0].value, rows[i + 1][1].value

    ew.save(file_path)
    return


# 规定读取中文xml 文件,返回中文字典，关键词顺序列表
def read_xml_file(file_path):
    global CN_map, EN_map, CN_key_list, EN_key_list
    cleanBuffer()
    xml_tree = parse(file_path)
    resources = xml_tree.documentElement
    StringKeyMaps = resources.getElementsByTagName('string')

    for item in StringKeyMaps:
        key = item.getAttribute('name')
        if key == None:
            continue
        CN_key_list.append(key)
        value = item.childNodes[0].nodeValue
        CN_map[key] = value
        # print key, '  :  ', value, '\n'

    return CN_map, CN_key_list


# 将翻译的xlsx文件写入xml中，如果有英文，写入相应的value,如果英文列处为空，写入相应的中文
def write_xml_file(EN_content, keys, default_content, file_path):
    if os.path.exists(file_path) == False:
        createFile(file_path)

    imp = minidom.getDOMImplementation()
    dom = imp.createDocument(None, None, None)
    root = dom.createElement('resources')

    root.setAttribute('xmlns:tools', 'http://schemas.android.com/tools')
    root.setAttribute('xmlns:xliff', 'urn:oasis:names:tc:xliff:document:1.2')
    root.setAttribute('tools:ignore', 'MissingTranslation')

    # for key, value in content.items():
    for key in keys:
        value = EN_content[key]
        if value == None:
            value = default_content[key]
        item = dom.createElement('string')
        item_value = dom.createTextNode(str(value))
        item.setAttribute('name', key)
        item.appendChild(item_value)

        root.appendChild(item)

        continue
    dom.appendChild(root)

    file_handle = open(file_path, 'w')
    dom.writexml(file_handle, '', '', '\n', 'utf-8')
    file_handle.close()

    return


def generate_android_strings_file(EN_content, key_list, CN_content):
    write_xml_file(EN_content, key_list, CN_content, GENERATE_ANDROID_STRINGS_FILE)
    return


def generate_PM_xlsx_file(content, key_list):
    write_to_xlsx_file(content, key_list, FILE_PATH_XLSX)
    return


def cleanBuffer():
    global EN_map, CN_map, CN_key_list, EN_key_list
    EN_map.clear()
    CN_map.clear()
    CN_key_list = []
    EN_key_list = []
    return


def createFile(file_name):
    f = open(file_name, 'w')
    f.close()
    return


# content = {'one': '1', 'two': '2', 'three': '3'}

if (__name__ == "__main__"):
    print 'this is cn_map', CN_map
    print 'this is cn_key_list', CN_key_list

    print sys.getdefaultencoding()

    # read_xml_file(FILE_PATH_XML)
    # generate_PM_xlsx_file(CN_map, CN_key_list)
    #
    read_xlsx_file(FILE_PATH_XLSX)
    generate_android_strings_file(EN_map, CN_key_list, CN_map)

    print 'succeed'
